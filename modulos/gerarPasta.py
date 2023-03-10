import shutil, os, webbrowser
from modulos.particao3Eixo import *
from modulos.fracionamentoX import *
from modulos.fracionamentoZ import *
from modulos.CriarPasta import *
import modulos.posicaoTabela as pt
from tkinter.filedialog import  askdirectory
from openpyxl import load_workbook
from tkinter import Tk

Tk().withdraw() 

unidadeDeDisco = str(os.path.abspath(os.sep))
usuario_atual_win = str(os.getlogin())


# ------------------------------------------------------------------------------------------------------------------------
# Recolhendo os valores das coordenadas X, Y e Z

def gerar_pastas(nome_da_amostra, volume_total):
    
    local_salvamento = askdirectory()

    metodo1 = particao_em_3eixos(volume_total)
    metodo2 = fracionamento_em_x(volume_total)
    metodo3 = fracionamento_em_z(volume_total)

    CriarSubamostras(local_salvamento, nome_da_amostra)
    webbrowser.open(os.path.realpath(f'{local_salvamento}\Amostra {nome_da_amostra}'))

# ------------------------------------------------------------------------------------------------------------------------
# importando os dados para a planilha método 01

    wb = load_workbook(fr'{os.getcwd()}\analise-rev\modulos\Modelo.xlsx')
    ws = wb['Plan1']

    Tabela01 = pt.celulasTabela01()
    c=0
    for d in range (0,15):
        for k in range (0, 3):
            ws[Tabela01[c]] = str(metodo1[d][k])
            c += 1

    wb.save(f'{os.getcwd()}\Subvolume da amostra X.xlsx')

# ------------------------------------------------------------------------------------------------------------------------
# importando os dados para a planilha método 02 em X

    wb = load_workbook(f'{os.getcwd()}\Subvolume da amostra X.xlsx')
    ws = wb['Plan1']

    Tabela02 = pt.celulasTabela02()
    c=0
    for d in range (0,14):
        for k in range (0, 3):
            ws[Tabela02[c]] = str(metodo2[d][k])
            c += 1
            
    wb.save(f'{os.getcwd()}\Subvolume da amostra X.xlsx')
# ------------------------------------------------------------------------------------------------------------------------
# importando os dados para a planilha método 03 em Z

    wb = load_workbook(f'{os.getcwd()}\Subvolume da amostra X.xlsx')
    ws = wb['Plan1']

    Tabela04 = pt.celulasTabela04()
    c=0
    for d in range (0,14):
        for k in range (0, 3):
            ws[Tabela04[c]] = str(metodo3[d][k])
            c += 1

# ------------------------------------------------------------------------------------------------------------------------
# Copiando o arqivo para área de trabalho
    wb.save(f'{os.getcwd()}\Subvolume da amostra X.xlsx')
    source = rf'{os.getcwd()}\Subvolume da amostra X.xlsx'
    destination = rf'{local_salvamento}\Amostra {nome_da_amostra}'
    shutil.copy(source, destination)
    os.rename(rf'{local_salvamento}\Amostra {nome_da_amostra}\Subvolume da amostra X.xlsx', rf'{local_salvamento}\Amostra {nome_da_amostra}\Subvolume da amostra {nome_da_amostra}.xlsx')

# ------------------------------------------------------------------------------------------------------------------------


