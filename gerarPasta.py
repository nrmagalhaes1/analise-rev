import particao3Eixo, shutil, fracionamentoX,fracionamentoZ, CriarPasta, os, webbrowser
from tkinter.filedialog import askopenfilename, askdirectory
from openpyxl import load_workbook
from tkinter import Tk
import formatacaoDeDados as f
import posicaoTabela as pt
Tk().withdraw() 

unidadeDeDisco = str(os.path.abspath(os.sep))
users = str(os.getlogin())


# ------------------------------------------------------------------------------------------------------------------------
# Recolhendo os valores das coordenadas X, Y e Z

def gerar(amostra, volume):
    print( 'Selecione um diretório para criar as pastas')
    diretorioSalve = askdirectory()
    Tk().withdraw() 

    metodo1 = particao3Eixo.executar(volume)
    metodo2 = fracionamentoX.executar(volume)
    metodo3 = fracionamentoZ.executar(volume)
    f.pularLin()

    CriarPasta.CriarSubamostras(diretorioSalve, amostra)
    webbrowser.open(os.path.realpath(f'{diretorioSalve}\Amostra {amostra}'))

# ------------------------------------------------------------------------------------------------------------------------
# importando os dados para a planilha método 01

    wb = load_workbook(f'{os.getcwd()}\Modelo.xlsx')
    ws = wb['Plan1']

    Tabela01 = pt.celulasTabela01()
    c=0
    for d in range (0,15):
        for k in range (0, 3):
            ws[Tabela01[c]] = str(metodo1[d][k])
            c += 1

    wb.save(f'{os.getcwd()}\Subvolume da amostra X.xlsx')

# ------------------------------------------------------------------------------------------------------------------------
# importando os dados para a planilha método 02 em X

    wb = load_workbook(f'{os.getcwd()}\Subvolume da amostra X.xlsx')
    ws = wb['Plan1']

    Tabela02 = pt.celulasTabela02()
    c=0
    for d in range (0,14):
        for k in range (0, 3):
            ws[Tabela02[c]] = str(metodo2[d][k])
            c += 1
            
    wb.save(f'{os.getcwd()}\Subvolume da amostra X.xlsx')
# ------------------------------------------------------------------------------------------------------------------------
# importando os dados para a planilha método 03 em Z

    wb = load_workbook(f'{os.getcwd()}\Subvolume da amostra X.xlsx')
    ws = wb['Plan1']

    Tabela04 = pt.celulasTabela04()
    c=0
    for d in range (0,14):
        for k in range (0, 3):
            ws[Tabela04[c]] = str(metodo3[d][k])
            c += 1

# ------------------------------------------------------------------------------------------------------------------------
# Copiando o arqivo para área de trabalho
    wb.save(f'{os.getcwd()}\Subvolume da amostra X.xlsx')
    source = rf'{os.getcwd()}\Subvolume da amostra X.xlsx'
    destination = rf'{diretorioSalve}\Amostra {amostra}'
    shutil.copy(source, destination)
    os.rename(rf'{diretorioSalve}\Amostra {amostra}\Subvolume da amostra X.xlsx', rf'{diretorioSalve}\Amostra {amostra}\Subvolume da amostra {amostra}.xlsx')

# ------------------------------------------------------------------------------------------------------------------------


