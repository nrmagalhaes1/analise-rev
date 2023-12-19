# Módulo para geração de subamostras a partir de uma amostra principal
import shutil  # Módulo para operações em arquivos e pastas
import os  # Módulo para interações com o sistema operacional
import webbrowser  # Módulo para controle do navegador
from Particao_3_eixos import *  # Módulo personalizado para particionamento em 3 eixos
from fracionamento_3_eixos import *  # Módulo personalizado para fracionamento em 3 eixos
from Criar_pasta import *  # Módulo personalizado para criação de pastas
import posicao_tabela as pt  # Módulo personalizado para posição em tabela
from tkinter.filedialog import askdirectory  # Diálogo de seleção de diretório
from openpyxl import load_workbook  # Manipulação de arquivos xlsx
from tkinter import Tk  # Interface gráfica

Tk().withdraw()  # Inicialização da interface gráfica e ocultação da janela

unidadeDeDisco = str(os.path.abspath(os.sep))  # Obter unidade de disco
usuario_atual_win = str(os.getlogin())  # Obter usuário atual do Windows

# Função para gerar pastas de subamostras a partir de uma amostra principal
def gerar_pastas(nome_da_amostra, volume_total):
    
    local_salvamento = askdirectory()  # Seleção do diretório de salvamento

    # Utilização de métodos de particionamento e fracionamento
    metodo1 = particao_em_3eixos(volume_total)
    metodo2 = fracionamento(volume_total)

    CriarSubamostras(local_salvamento, nome_da_amostra)  # Criação de subamostras
    webbrowser.open(os.path.realpath(f'{local_salvamento}\Amostra {nome_da_amostra}'))  # Abertura do diretório no navegador

    # Manipulação de dados de particionamento e fracionamento
    lista = particao_em_3eixos(volume_total) + fracionamento(volume_total)
    q, w, e, r, t, y = [], [], [], [], [], []

    # Iteração para categorização dos dados
    for vetor in lista:
        if isinstance(vetor[0], list):
            q.append(int(vetor[0][0]))
            r.append(int(vetor[0][1]))
        else:
            q.append(int(1))
            r.append(int(vetor[0]))

        # (continuação similar para 'w', 't', 'e', 'y')

    # Tratamento de valores zero para evitar erros
    q = [1 if x == 0 else x for x in q]
    w = [1 if x == 0 else x for x in w]
    # (continuação para 'e', 'r', 't', 'y')

    # Exibição dos resultados
    print(lista)
    print("\n")
    # (continuação para exibição de 'q', 'w', 'e', 'r', 't', 'y')

    # Carregamento de um modelo de arquivo xlsx e preenchimento com dados do método 1
    wb = load_workbook(fr'{os.getcwd()}\modulos\Modelo.xlsx')
    ws = wb['Plan1']
    Tabela01 = pt.celulasTabela01()

    # Preenchimento das células da Tabela01 com os dados do método 1
    c = 0
    for d in range(0, 15):
        for k in range(0, 3):
            ws[Tabela01[c]] = str(metodo1[d][k])
            c += 1

    # Salvamento do arquivo com os dados do método 1
    wb.save(f'{os.getcwd()}\Subvolume da amostra X.xlsx')

    # Carregamento do arquivo salvo e preenchimento com dados do método 2
    wb = load_workbook(f'{os.getcwd()}\Subvolume da amostra X.xlsx')
    ws = wb['Plan1']
    Tabela02 = pt.celulasTabela02()

    # Preenchimento das células da Tabela02 com os dados do método 2
    c = 0
    for d in range(0, 23):
        for k in range(0, 3):
            ws[Tabela02[c]] = str(metodo2[d][k])
            c += 1
            
    # Salvamento do arquivo com os dados do método 2
    wb.save(f'{os.getcwd()}\Subvolume da amostra X.xlsx')

    # Cópia e renomeação do arquivo para o diretório de destino
    wb.save(f'{os.getcwd()}\Subvolume da amostra X.xlsx')
    source = rf'{os.getcwd()}\Subvolume da amostra X.xlsx'
    destination = rf'{local_salvamento}\Amostra {nome_da_amostra}'
    shutil.copy(source, destination)
    os.rename(rf'{local_salvamento}\Amostra {nome_da_amostra}\Subvolume da amostra X.xlsx', rf'{local_salvamento}\Amostra {nome_da_amostra}\Subvolume da amostra {nome_da_amostra}.xlsx')
