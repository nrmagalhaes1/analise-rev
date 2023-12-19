# Módulo de Análise de Dados em Planilhas Excel
# Este módulo fornece funções para extrair informações específicas de planilhas Excel.
# Desenvolvido para analisar dados em planilhas contendo valores de phi e volume.

# Bibliotecas necessárias
import openpyxl  # Biblioteca para trabalhar com planilhas Excel
import os        # Biblioteca para interagir com o sistema operacional

# Função para recuperar os valores de phi de uma planilha Excel
def recolherValorPhi(filename):
    """
    :param filename: O caminho do arquivo da planilha Excel.
    :return: Uma lista contendo os valores de phi extraídos da planilha.
    """
    path = rf'{filename}'
    wb_obj = openpyxl.load_workbook(path, data_only=True) 
    sheet_obj = wb_obj.active

    phi = []
    c = 0
    # Extrair valores de phi da primeira seção da planilha
    for i in range(8, 22):
        cell_obj = sheet_obj.cell(row=i, column=7) 
        phi.append(cell_obj.value)
        c += 1

    c = 0
    # Extrair valores de phi da segunda seção da planilha
    for i in range(25, 48):
        cell_obj = sheet_obj.cell(row=i, column=7) 
        phi.append(cell_obj.value)
        c += 1

    return phi

# Função para recuperar os valores de volume de uma planilha Excel
def recolherValorvolume(filename):
    """
    :param filename: O caminho do arquivo da planilha Excel.
    :return: Uma lista contendo os valores de volume extraídos da planilha.
    """
    path = rf'{filename}'
    wb_obj = openpyxl.load_workbook(path, data_only=True) 
    sheet_obj = wb_obj.active

    volume = []
    c = 0
    # Extrair valores de volume da primeira seção da planilha
    for i in range(8, 22):
        cell_obj = sheet_obj.cell(row=i, column=11)
        volume.append(cell_obj.value)
        c += 1

    c = 0
    # Extrair valores de volume da segunda seção da planilha
    for i in range(25, 48):
        cell_obj = sheet_obj.cell(row=i, column=11) 
        volume.append(cell_obj.value)
        c += 1

    return volume

# Função para recuperar os valores totais de uma planilha Excel
def valor_total(filename):
    """
    :param filename: O caminho do arquivo da planilha Excel.
    :return: Uma lista contendo os valores totais extraídos da planilha.
    """
    total = []
    path = rf'{filename}'
    wb_obj = openpyxl.load_workbook(path, data_only=True) 
    sheet_obj = wb_obj.active
    # Extrair valores totais das células específicas da planilha
    cell_obj = sheet_obj.cell(row=7, column=7) 
    total.append(cell_obj.value)
    cell_obj = sheet_obj.cell(row=7, column=11) 
    total.append(cell_obj.value)

    return total
